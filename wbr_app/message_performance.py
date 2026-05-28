from fastapi import FastAPI, Query
from fastapi.responses import JSONResponse
from urllib.parse import quote
from fastapi.responses import HTMLResponse, StreamingResponse
from google.cloud import bigquery
from datetime import datetime, timedelta
from collections import OrderedDict
from io import BytesIO
import json

app    = FastAPI(title='HP Message Performance')
client = bigquery.Client()

FYTD_START = '2026-01-31'
BQ_HSA     = 'wmt-site-content-strategy.scs_production.hp_summary_asset'
BQ_HPS     = 'wmt-site-content-strategy.scs_production.HPsummary'
SIG_CARDS  = "('SIG Card 1','SIG Card 2','SIG Card 3','SIG Card 4','SIG Card 5','SIG Card 6')"
HPOV_CARDS = "('AutoScroll Card 1','AutoScroll Card 2','AutoScroll Card 3','AutoScroll Card 4','AutoScroll Card 5')"
DEFAULT_SERVICES = 'Get It Fast|Walmart Plus'

CT_HSA = """
AND CASE
  WHEN LOWER(IFNULL(content_served_by,'')) = 'ads' THEN 'WMC'
  WHEN LOWER(IFNULL(disable_content_personalization,'')) LIKE '%true%' THEN 'Merch'
  WHEN LOWER(IFNULL(disable_content_personalization,'')) LIKE '%false%'
    AND LOWER(IFNULL(personalized_asset,'')) = 'default'
    AND session_start_dt <= '2025-03-01'
    AND LOWER(IFNULL(content_zone,'')) = 'contentzone3'
    AND LOWER(IFNULL(hp_module_name,'')) IN (
      'autoscroll card 1','autoscroll card 2','autoscroll card 3') THEN 'WMC'
  WHEN LOWER(IFNULL(disable_content_personalization,'')) LIKE '%false%'
    AND LOWER(IFNULL(personalized_asset,'')) = 'default'
    AND ((LOWER(IFNULL(content_zone,'')) IN ('contentzone8','contentzone9')
          AND LOWER(IFNULL(hp_module_name,'')) = 'adjustable banner small')
      OR (LOWER(IFNULL(content_zone,'')) IN ('contentzone10','contentzone11')
          AND LOWER(IFNULL(hp_module_name,'')) = 'triple pack small')) THEN 'WMC'
  ELSE 'Merch'
END = 'Merch'
"""

PALETTE = [
    '#0071CE','#22c55e','#E63946','#f59e0b','#8b5cf6','#ec4899',
    '#14b8a6','#06b6d4','#6366f1','#a855f7','#f97316','#84cc16',
    '#0ea5e9','#d946ef','#64748b','#FFC220'
]

# ─── HPOV DATA ─────────────────────────────────────────────────────────
# Filters only by hp_module_name — no content_zone filter
# MM SIG launch May 14 2026 moved HPOV from contentZone3 → contentZone2

def get_hpov_data(start, end, language='English', services_m0=None):
    svc_list = services_m0 or ['Get It Fast', 'Walmart Plus']
    svc_sql  = ','.join(f"'{s.replace(chr(39), chr(39)+chr(39))}'" for s in svc_list)
    lf       = f"AND lang_cd = '{language}'" if language != 'All' else ''
    fytd_end = f"""(SELECT MAX(session_start_dt) FROM `{BQ_HSA}`
                    WHERE session_start_dt >= '{FYTD_START}')"""
    q = f"""
    WITH
    fytd_bench AS (
      SELECT SAFE_DIVIDE(SUM(overall_click_count), SUM(module_view_count)) * 100 AS bm
      FROM `{BQ_HSA}`
      WHERE session_start_dt BETWEEN '{FYTD_START}' AND {fytd_end}
        AND hp_module_name IN {HPOV_CARDS}
        AND experience_lvl2 IN ('App: iOS', 'App: Android') {lf} {CT_HSA}
    ),
    curr AS (
      SELECT
        CASE WHEN COALESCE(m0_nm,'') IN ({svc_sql}) THEN 'Services'
             ELSE COALESCE(NULLIF(m0_nm,''), 'Unknown') END AS m0_group,
        COALESCE(NULLIF(message_name,''), 'Unknown') AS msg,
        SUM(module_view_count) AS imp,
        SUM(overall_click_count) AS clk,
        SUM(total_atc_count) AS atc_cnt,
        SAFE_DIVIDE(SUM(overall_click_count), SUM(module_view_count)) * 100 AS ctr,
        SAFE_DIVIDE(SUM(total_atc_count), SUM(module_view_count)) * 1000 AS atc_rate
      FROM `{BQ_HSA}`
      WHERE session_start_dt BETWEEN '{start}' AND '{end}'
        AND hp_module_name IN {HPOV_CARDS}
        AND experience_lvl2 IN ('App: iOS', 'App: Android')
        AND message_name IS NOT NULL AND m0_nm IS NOT NULL
        {lf} {CT_HSA}
      GROUP BY m0_group, msg HAVING SUM(module_view_count) > 100000
    ),
    m0t AS (SELECT m0_group, SUM(imp) AS m0_imp FROM curr GROUP BY m0_group),
    gt  AS (SELECT SUM(imp) AS total FROM curr)
    SELECT c.m0_group, c.msg AS message_name,
      ROUND(c.imp / 1000000, 2) AS imp_m, ROUND(c.ctr, 2) AS ctr,
      ROUND(c.atc_rate, 2) AS atc_rate,
      ROUND(c.imp  / g.total * 100, 1) AS sov_pct,
      ROUND(m0t.m0_imp / g.total * 100, 1) AS m0_sov,
      ROUND(b.bm, 2) AS fytd_baseline
    FROM curr c CROSS JOIN gt g CROSS JOIN fytd_bench b
    LEFT JOIN m0t ON c.m0_group = m0t.m0_group
    ORDER BY m0t.m0_imp DESC, c.imp DESC
    """
    try: return [dict(r) for r in client.query(q).result()]
    except Exception as e: print(f'HPOV err: {e}'); return []


def get_hpov_m0_list(start, end, language='English'):
    """All distinct m0_nm values for the Services selector."""
    lf = f"AND lang_cd = '{language}'" if language != 'All' else ''
    q  = f"""
    SELECT COALESCE(NULLIF(m0_nm,''),'Unknown') AS m0_nm,
           SUM(module_view_count) AS imp
    FROM `{BQ_HSA}`
    WHERE session_start_dt BETWEEN '{start}' AND '{end}'
      AND hp_module_name IN {HPOV_CARDS}
      AND experience_lvl2 IN ('App: iOS','App: Android')
      AND m0_nm IS NOT NULL {lf} {CT_HSA}
    GROUP BY 1 HAVING SUM(module_view_count) > 100000
    ORDER BY imp DESC
    """
    try: return [dict(r) for r in client.query(q).result()]
    except Exception as e: print(f'M0 list err: {e}'); return []


def get_hpov_messages(start, end, language='English', services_m0=None):
    """All distinct message_name values for HPOV — for the message filter selector."""
    svc_list = services_m0 or ['Get It Fast', 'Walmart Plus']
    svc_sql  = ','.join(f"'{s.replace(chr(39), chr(39)+chr(39))}'" for s in svc_list)
    lf = f"AND lang_cd = '{language}'" if language != 'All' else ''
    q  = f"""
    SELECT
      CASE WHEN COALESCE(m0_nm,'') IN ({svc_sql}) THEN 'Services'
           ELSE COALESCE(NULLIF(m0_nm,''),'Unknown') END AS m0_group,
      COALESCE(NULLIF(message_name,''),'Unknown') AS message_name,
      SUM(module_view_count) AS imp
    FROM `{BQ_HSA}`
    WHERE session_start_dt BETWEEN '{start}' AND '{end}'
      AND hp_module_name IN {HPOV_CARDS}
      AND experience_lvl2 IN ('App: iOS','App: Android')
      AND message_name IS NOT NULL AND m0_nm IS NOT NULL
      {lf} {CT_HSA}
    GROUP BY 1,2 HAVING SUM(module_view_count) > 100000
    ORDER BY 1, imp DESC
    """
    try: return [dict(r) for r in client.query(q).result()]
    except Exception as e: print(f'HPOV msg err: {e}'); return []


def get_sig_carousels(start, end, language='English'):
    lf = f"AND lang_cd = '{language}'" if language != 'All' else ''
    q  = f"""
    SELECT COALESCE(NULLIF(carousel_name,''),'Unknown') AS carousel_nm,
           SUM(module_view_count) AS imp
    FROM `{BQ_HSA}`
    WHERE session_start_dt BETWEEN '{start}' AND '{end}'
      AND hp_module_name IN {SIG_CARDS} {lf} {CT_HSA}
    GROUP BY 1 HAVING SUM(module_view_count) > 500000
    ORDER BY imp DESC
    """
    try: return [dict(r) for r in client.query(q).result()]
    except Exception as e: print(f'SIG car err: {e}'); return []


def get_sig_messages(start, end, language='English', selected_carousels=None):
    lf  = f"AND lang_cd = '{language}'" if language != 'All' else ''
    caf = ''
    if selected_carousels:
        esc = ','.join(f"'{c.replace(chr(39),chr(39)+chr(39))}'" for c in selected_carousels)
        caf = f'AND carousel_name IN ({esc})'
    q = f"""
    SELECT COALESCE(NULLIF(message_name,''), carousel_name, 'Unknown') AS display_label,
           SUM(module_view_count) AS imp
    FROM `{BQ_HSA}`
    WHERE session_start_dt BETWEEN '{start}' AND '{end}'
      AND hp_module_name IN {SIG_CARDS}
      AND NOT (LOWER(COALESCE(carousel_name,'')) LIKE '%top pick%'
           OR  LOWER(COALESCE(carousel_name,'')) LIKE '%jump right back%')
      {lf} {caf} {CT_HSA}
    GROUP BY 1 HAVING SUM(module_view_count) > 50000
    ORDER BY imp DESC
    """
    try: return [dict(r) for r in client.query(q).result()]
    except Exception as e: print(f'SIG msg err: {e}'); return []


def get_sig_data(start, end, language='English', selected_carousels=None, selected_messages=None):
    lf  = f"AND lang_cd = '{language}'" if language != 'All' else ''
    caf = ''
    if selected_carousels:
        esc = ','.join(f"'{c.replace(chr(39),chr(39)+chr(39))}'" for c in selected_carousels)
        caf = f'AND carousel_name IN ({esc})'
    fytd_end = f"""(SELECT MAX(session_start_dt) FROM `{BQ_HSA}`
                    WHERE session_start_dt >= '{FYTD_START}')"""
    q = f"""
    WITH
    fytd_p13n AS (
      SELECT SAFE_DIVIDE(SUM(overall_click_count), SUM(module_view_count)) * 100 AS bm
      FROM `{BQ_HSA}`
      WHERE session_start_dt BETWEEN '{FYTD_START}' AND {fytd_end}
        AND hp_module_name IN {SIG_CARDS}
        AND (LOWER(COALESCE(carousel_name,'')) LIKE '%top pick%'
          OR LOWER(COALESCE(carousel_name,'')) LIKE '%jump right back%')
        {lf} {CT_HSA}
    ),
    fytd_merch AS (
      SELECT SAFE_DIVIDE(SUM(overall_click_count), SUM(module_view_count)) * 100 AS bm
      FROM `{BQ_HSA}`
      WHERE session_start_dt BETWEEN '{FYTD_START}' AND {fytd_end}
        AND hp_module_name IN {SIG_CARDS}
        AND NOT (LOWER(COALESCE(carousel_name,'')) LIKE '%top pick%'
              OR LOWER(COALESCE(carousel_name,'')) LIKE '%jump right back%')
        {lf} {CT_HSA}
    ),
    curr AS (
      SELECT
        COALESCE(NULLIF(carousel_name,''),'Unknown') AS carousel_nm,
        CASE WHEN LOWER(COALESCE(carousel_name,'')) LIKE '%top pick%'
               OR LOWER(COALESCE(carousel_name,'')) LIKE '%jump right back%'
             THEN 'P13N' ELSE 'Site Merch' END AS carousel_type,
        CASE WHEN LOWER(COALESCE(carousel_name,'')) LIKE '%top pick%'
               OR LOWER(COALESCE(carousel_name,'')) LIKE '%jump right back%'
             THEN COALESCE(NULLIF(asset_heading,''), message_name, 'Unknown')
             ELSE COALESCE(NULLIF(message_name,''), carousel_name, 'Unknown')
        END AS display_label,
        SUM(module_view_count) AS imp,
        SUM(overall_click_count) AS clk,
        SAFE_DIVIDE(SUM(overall_click_count), SUM(module_view_count)) * 100 AS ctr
      FROM `{BQ_HSA}`
      WHERE session_start_dt BETWEEN '{start}' AND '{end}'
        AND hp_module_name IN {SIG_CARDS} {lf} {caf} {CT_HSA}
      GROUP BY carousel_nm, carousel_type, display_label
      HAVING SUM(module_view_count) > 100000
    ),
    ct  AS (SELECT carousel_nm, SUM(imp) AS c_imp FROM curr GROUP BY carousel_nm),
    gt  AS (SELECT SUM(imp) AS total FROM curr)
    SELECT c.carousel_nm, c.carousel_type, c.display_label, c.imp, c.clk,
      ROUND(c.imp / 1000000, 2) AS imp_m, ROUND(c.ctr, 2) AS ctr,
      ROUND(c.atc_rate, 2) AS atc_rate,
      ROUND(c.imp  / g.total * 100, 1) AS sov_pct,
      ROUND(ct.c_imp / g.total * 100, 1) AS carousel_sov,
      ROUND(bp.bm, 2) AS fytd_baseline_p13n,
      ROUND(bm.bm, 2) AS fytd_baseline_merch
    FROM curr c CROSS JOIN gt g
    CROSS JOIN fytd_p13n bp CROSS JOIN fytd_merch bm
    LEFT JOIN ct ON c.carousel_nm = ct.carousel_nm
    ORDER BY c.carousel_type DESC, ct.c_imp DESC, c.imp DESC
    """
    try:
        rows = [dict(r) for r in client.query(q).result()]
        if selected_messages:
            sm = set(selected_messages)
            rows = [r for r in rows if r.get('carousel_type') == 'P13N'
                    or r.get('display_label') in sm]
        return rows
    except Exception as e: print(f'SIG err: {e}'); return []


def _bar_chart(cid, groups, baseline):
    labels,imp_d,ctr_d,bc,bands=[],[],[],[],[]
    for g in groups:
        bands.append({'n':g['name'],'cnt':len(g['msgs']),'sov':g['sov'],'col':g['color']})
        for m in g['msgs']:
            labels.append(m['lbl']); imp_d.append(float(m['imp']))
            ctr_d.append(float(m['ctr'])); bc.append(g['color'])
    lj=json.dumps(labels); ij=json.dumps(imp_d); cj=json.dumps(ctr_d)
    bj=json.dumps(bc); bdj=json.dumps(bands); bm=baseline
    return (
        "<div style='background:white;border-radius:10px;box-shadow:0 2px 8px rgba(0,0,0,.1);"
        "padding:20px;margin-bottom:24px;'>"
        f"<div style='height:440px;'><canvas id='{cid}'></canvas></div>"
        f"<div id='{cid}Band' style='display:flex;margin:0 50px;'></div>"
        "</div><script>(function(){"
        f"var L={lj},I={ij},C={cj},BC={bj},BD={bdj},BM={bm};"
        f"var ctx=document.getElementById('{cid}').getContext('2d');"
        "new Chart(ctx,{type:'bar',data:{labels:L,datasets:["
        "{label:'Imps(M)',data:I,backgroundColor:BC,borderRadius:4,order:2,yAxisID:'y'},"
        "{label:'CTR%',data:C,type:'line',borderColor:'#1f2937',backgroundColor:'#1f2937',"
        "pointRadius:5,pointBackgroundColor:'#fff',pointBorderColor:'#1f2937',"
        "pointBorderWidth:2,tension:0,order:1,yAxisID:'y1'}"
        "]},options:{responsive:true,maintainAspectRatio:false,"
        "plugins:{legend:{display:false},"
        "datalabels:{display:true,"
        "anchor:function(c){return c.datasetIndex===0?'center':'end';},"
        "align:function(c){return c.datasetIndex===0?'center':'top';},"
        "color:function(c){return c.datasetIndex===0?'#fff':'#1f2937';},"
        "font:{weight:'bold',size:9},"
        "formatter:function(v,c){return c.datasetIndex===0?v+'M':(+v).toFixed(2)+'%';}},"
        "annotation:{annotations:{bl:{type:'line',yMin:BM,yMax:BM,yScaleID:'y1',"
        "borderColor:'#FFC220',borderWidth:2.5,borderDash:[6,4],"
        "label:{display:true,content:'Benchmark '+BM.toFixed(2)+'%',position:'end',"
        "backgroundColor:'#FFC220',color:'#1a1a2e',font:{size:10,weight:'bold'},padding:3}}}}},"
        "scales:{x:{grid:{display:false},ticks:{maxRotation:45,minRotation:45,font:{size:8}}},"
        "y:{type:'linear',position:'left',grid:{color:'#e5e7eb'},min:0},"
        "y1:{type:'linear',position:'right',grid:{display:false},min:0}}},"
        "plugins:[ChartDataLabels]});"
        f"var band=document.getElementById('{cid}Band');"
        "BD.forEach(function(b){var d=document.createElement('div');"
        "d.style.cssText='display:flex;align-items:center;justify-content:center;"
        "color:white;font-weight:600;font-size:10px;text-align:center;padding:6px 2px;"
        "line-height:1.2;background-color:'+b.col+';flex:'+b.cnt;"
        "d.innerHTML=b.n+'<br>'+b.sov+'% SOV';band.appendChild(d);});"
        "})();</script>"
    )


def render_hpov_chart(data):
    if not data: return '<p style="color:#888;">No HPOV data.</p>'
    bl = float(data[0].get('fytd_baseline') or 0.21)
    gs,ci = OrderedDict(),0
    for r in data:
        g = r['m0_group']
        if g not in gs:
            gs[g]={'name':g,'color':PALETTE[ci%len(PALETTE)],'sov':float(r.get('m0_sov') or 0),'msgs':[]}
            ci+=1
        gs[g]['msgs'].append({'lbl':r['message_name'],'imp':float(r.get('imp_m') or 0),'ctr':float(r.get('ctr') or 0)})
    return _bar_chart('hpovC', list(gs.values()), bl)


def render_sig_chart(data):
    if not data: return '<p style="color:#888;">No SIG data.</p>'
    bl = float(data[0].get('fytd_baseline') or 1.11)
    cs,ci = OrderedDict(),0
    for r in data:
        cn = r['carousel_nm']
        if cn not in cs:
            col='#3b82f6' if r['carousel_type']=='P13N' else PALETTE[ci%len(PALETTE)]
            cs[cn]={'name':cn,'color':col,'sov':float(r.get('carousel_sov') or 0),'msgs':[]}
            ci+=1
        cs[cn]['msgs'].append({'lbl':str(r.get('display_label') or '')[:20],'imp':float(r.get('imp_m') or 0),'ctr':float(r.get('ctr') or 0)})
    return _bar_chart('sigC', list(cs.values()), bl)


_BD  = 'border:1px solid #e5e7eb;'
_BDH = 'border:1px solid #cbd5e1;'
_TH  = 'style="padding:9px 12px;text-align:center;font-size:12px;font-weight:700;border:1px solid #1a3a60;white-space:nowrap;"'
_THL = 'style="padding:9px 12px;text-align:left;font-size:12px;font-weight:700;border:1px solid #1a3a60;white-space:nowrap;"'
_TD  = f'style="padding:7px 12px;text-align:center;{_BD}"'
_TDL = f'style="padding:7px 12px;text-align:left;{_BD}"'

VAR_JS = """
<script>
function calcVar(inp){
  var a=parseFloat(inp.dataset.actual)||0;
  var p=parseFloat(inp.value);
  var el=document.getElementById(inp.dataset.rid+'_var');
  if(!el||isNaN(p)){if(el){el.textContent='--';el.style.color='#999';}return;}
  var v=(a-p).toFixed(1);
  el.textContent=(parseFloat(v)>=0?'+':'')+v+'%';
  el.style.color=parseFloat(v)>=0?'#16a34a':'#dc2626';
  el.style.fontWeight='700';
}
</script>
"""

def _inp(rid, sov, msg=''):
    sm = msg.replace('"','&quot;')
    return (
        f'<input type="number" step="0.1" min="0" max="100" '
        f'data-rid="{rid}" data-actual="{sov}" data-msg="{sm}" oninput="calcVar(this)" '
        f'style="width:64px;border:1px solid #d1d5db;border-radius:4px;'
        f'padding:3px 5px;font-size:11px;text-align:center;">'
    )

def _var_td(rid):
    return f'<td id="{rid}_var" style="padding:7px 12px;text-align:center;{_BD}color:#999;">--</td>'

def _ctr_td(v, bl):
    c='#16a34a' if v>=bl else '#dc2626'
    return f'<td style="padding:7px 12px;text-align:center;{_BD}font-weight:700;color:{c};">{v:.2f}%</td>'

def _thead():
    """HPOV table header — 7 cols."""
    return (
        '<thead><tr style="background:#041e42;color:white;">'
        f'<th {_THL}>Group</th>'
        f'<th {_THL}>Message / Asset</th>'
        f'<th {_TH}>SOV Projected (%)</th>'
        f'<th {_TH}>SOV Actual (%)</th>'
        f'<th {_TH}>SOV Variance (+/-)</th>'
        f'<th {_TH}>CTR %</th>'
        f'<th {_TH}>ATC Rate</th>'
        '</tr></thead>'
    )


def _sig_thead():
    """SIG table header — 5 cols."""
    return (
        '<thead><tr style="background:#041e42;color:white;">'
        f'<th {_THL}>Group / Carousel</th>'
        f'<th {_THL}>Message / Asset</th>'
        f'<th {_TH}>SOV %</th>'
        f'<th {_TH}>CTR %</th>'
        f'<th {_TH}>ATC Rate</th>'
        '</tr></thead>'
    )


def _dd_panel(dd_id, title, options, selected_set, apply_fn, grouped=False):
    """
    Build a custom dropdown panel.
    options: list of (label, value, count) or (group, label, value, count) if grouped
    selected_set: set of currently selected values
    apply_fn: JS function name to call on Apply
    """
    # Count selected
    total = len(options)
    n_sel = len(selected_set) if selected_set else total
    badge = f'{n_sel} of {total}' if selected_set else 'All'

    # Build options HTML
    opts_html = ''
    if grouped:
        by_grp = {}
        for grp, lbl, val, cnt in options:
            if grp not in by_grp: by_grp[grp] = []
            by_grp[grp].append((lbl, val, cnt))
        for grp, items in by_grp.items():
            opts_html += f'<div class="dd-grp">{grp}</div>'
            for lbl, val, cnt in items:
                chk = 'checked' if (not selected_set or val in selected_set) else ''
                sv  = val.replace('"', '&quot;')
                opts_html += (
                    f'<label><input type="checkbox" class="dd-opt-{dd_id}" value="{sv}" {chk}>'
                    f' {lbl} <span style="color:#9ca3af;font-size:10px;">({cnt}M)</span></label>'
                )
    else:
        for lbl, val, cnt in options:
            chk = 'checked' if (not selected_set or val in selected_set) else ''
            sv  = val.replace('"', '&quot;')
            opts_html += (
                f'<label><input type="checkbox" class="dd-opt-{dd_id}" value="{sv}" {chk}>'
                f' {lbl} <span style="color:#9ca3af;font-size:10px;">({cnt}M)</span></label>'
            )

    html = (
        f'<div class="dd-wrap">'
        f'<button type="button" class="dd-btn" onclick="ddToggle(\'{dd_id}\')" '
        f'style="min-width:220px;">'
        f'<span>{title}</span>'
        f'<span style="background:#e5e7eb;border-radius:12px;padding:2px 8px;'
        f'font-size:11px;color:#374151;">{badge}</span>'
        f'<span style="color:#9ca3af;">&#9660;</span>'
        f'</button>'
        f'<div class="dd-panel" id="ddp_{dd_id}" style="display:none;">'
        f'<input type="text" class="dd-search" placeholder="Search..." '
        f'oninput="ddSearch(\'{dd_id}\', this.value)">'
        f'<div class="dd-opts" id="ddopts_{dd_id}">{opts_html}</div>'
        f'<div class="dd-footer">'
        f'<button type="button" class="dd-selall" onclick="ddSelAll(\'{dd_id}\',true)">All</button>'
        f'<button type="button" class="dd-clear" onclick="ddSelAll(\'{dd_id}\',false)">None</button>'
        f'<button type="button" class="dd-apply" onclick="{apply_fn}(\'{dd_id}\')">Apply &#10003;</button>'
        f'</div></div></div>'
    )
    return html


def render_hpov_table(data, sel_msgs=None):
    if not data: return '<p style="color:#888;">No HPOV data. Try a wider date range.</p>'
    bl = float(data[0].get('fytd_baseline') or 0.21)
    groups = OrderedDict()
    sm_set = set(sel_msgs) if sel_msgs else None
    for row in data:
        if sm_set and row.get('message_name') not in sm_set: continue
        g = row['m0_group']
        if g not in groups: groups[g] = []
        groups[g].append(row)
    # Baseline badge (top-right)
    badge = (
        f'<div style="display:flex;justify-content:flex-end;margin-bottom:8px;">'
        f'<span style="background:#041e42;color:white;font-size:12px;font-weight:700;'
        f'padding:6px 14px;border-radius:6px;">'
        f'FYTD CTR Baseline = {bl:.2f}%</span></div>'
    )
    html = (
        badge +
        f'<div style="overflow-x:auto;">'
        f'<table style="border-collapse:collapse;width:100%;font-size:12px;">'
        + _thead() + '<tbody>'
    )
    rn = 0
    for gname, rows in groups.items():
        n = len(rows)
        for j, row in enumerate(rows):
            bg  = '#f8f9fa' if rn % 2 else '#ffffff'
            ctr = float(row.get('ctr') or 0)
            sov = float(row.get('sov_pct') or 0)
            atc = float(row.get('atc_rate') or 0)
            rid = f'h{rn}'
            if j == 0:
                m0td = (
                    f'<td rowspan="{n}" style="padding:7px 12px;text-align:left;{_BD}'
                    f'font-weight:700;background:#f0f4ff;vertical-align:middle;font-size:11px;">'
                    f'{gname}</td>'
                )
            else:
                m0td = ''
            cc = '#16a34a' if ctr >= bl else '#dc2626'
            html += (
                f'<tr style="background:{bg};">{m0td}'
                f'<td {_TDL}>{row.get("message_name","—")}</td>'
                f'<td style="padding:7px 12px;text-align:center;{_BD}">{_inp(rid, sov, row.get("message_name",""))}</td>'
                + f'<td {_TD}>{sov:.1f}%</td>'
                + _var_td(rid)
                + f'<td style="padding:7px 12px;text-align:center;{_BD}font-weight:700;color:{cc};">{ctr:.2f}%</td>'
                + f'<td {_TD}>{atc:.1f}</td>'
                + '</tr>'
            )
            rn += 1
    html += f'</tbody></table></div>{VAR_JS}'
    return html


def render_sig_table(data, selected_messages=None):
    if not data: return '<p style="color:#888;">No SIG data. Check filters.</p>'

    def sf(r,k): return float(r.get(k) or 0)

    # ── Extract dual baselines ────────────────────────────────────────
    bl_p13n  = float(data[0].get('fytd_baseline_p13n')  or 0)
    bl_merch = float(data[0].get('fytd_baseline_merch') or 0)
    # Fallback: if baselines missing (old data), split the single baseline heuristically
    if bl_p13n == 0 and bl_merch == 0:
        fb = float(data[0].get('fytd_baseline') or 1.11)
        bl_p13n = fb; bl_merch = fb

    p13n  = [r for r in data if r.get('carousel_type') == 'P13N']
    merch = [r for r in data if r.get('carousel_type') == 'Site Merch']
    if selected_messages:
        sm = set(selected_messages)
        merch = [r for r in merch if r.get('display_label') in sm]

    def totals(rows):
        imp = sum(sf(r,'imp_m') for r in rows)
        clk = sum(sf(r,'imp_m')*sf(r,'ctr')/100 for r in rows)
        atc = sum(sf(r,'imp_m')*sf(r,'atc_rate') for r in rows)
        ctr_t = (clk/imp*100) if imp else 0
        atc_t = (atc/imp) if imp else 0
        return imp, ctr_t, atc_t

    tot_imp       = sum(sf(r,'imp_m') for r in data)
    p_imp, p_ctr, p_atc  = totals(p13n)
    m_imp, m_ctr, m_atc  = totals(merch)
    p_sov = round(p_imp/tot_imp*100,1) if tot_imp else 0
    m_sov = round(m_imp/tot_imp*100,1) if tot_imp else 0

    merch_cars = OrderedDict()
    for r in merch:
        cn = r['carousel_nm']
        if cn not in merch_cars: merch_cars[cn] = []
        merch_cars[cn].append(r)

    # 5-column total row — accepts its own baseline + atc_rate
    def total_row(label, sov, ctr, atc, bl):
        c = '#16a34a' if ctr >= bl else '#dc2626'
        return (
            f'<tr style="background:#f8fafc;font-weight:700;border-top:2px solid #cbd5e1;border-bottom:2px solid #cbd5e1;">'
            f'<td style="padding:10px 14px;border:1px solid #cbd5e1;font-size:13px;font-weight:800;border-left:4px solid #0071CE;">{label}</td>'
            f'<td style="padding:10px 14px;border:1px solid #cbd5e1;"></td>'
            f'<td style="padding:10px 14px;text-align:center;border:1px solid #cbd5e1;font-weight:800;">{sov:.1f}%</td>'
            f'<td style="padding:10px 14px;text-align:center;border:1px solid #cbd5e1;font-weight:800;color:{c};">{ctr:.2f}%</td>'
            f'<td style="padding:10px 14px;text-align:center;border:1px solid #cbd5e1;font-weight:800;">{atc:.1f}</td>'
            f'</tr>'
        )

    # Table (no info bar — baselines shown as badges after table)
    html = (
        f'<div style="overflow-x:auto;">'
        f'<table style="border-collapse:collapse;width:100%;font-size:12px;">'
        + _sig_thead() + '<tbody>'
    )
    rn = 0

    # ── P13N section ──────────────────────────────────────────────────
    if p13n:
        p_atc = (sum(sf(r,'imp_m')*sf(r,'atc_rate') for r in p13n)/p_imp) if p_imp else 0
        html += total_row('⬡ P13N — Top Picks for You', p_sov, p_ctr, p_atc, bl_p13n)
        rn += 1
        for row in p13n:
            bg  = '#f8f9fa' if rn % 2 else '#ffffff'
            ctr = sf(row,'ctr'); imp = sf(row,'imp_m'); sov = sf(row,'sov_pct')
            cc  = '#16a34a' if ctr >= bl_p13n else '#dc2626'
            atc = sf(row,'atc_rate')
            html += (
                f'<tr style="background:{bg};">'
                f'<td style="padding:7px 12px 7px 20px;{_BD}">'
                f'<span style="background:#3b82f6;color:white;border-radius:3px;'
                f'padding:1px 5px;font-size:9px;">P13N</span></td>'
                f'<td {_TDL}>{row.get("display_label","—")}</td>'
                f'<td {_TD}>{sov:.1f}%</td>'
                f'<td style="padding:7px 12px;text-align:center;{_BD}font-weight:700;color:{cc};">{ctr:.2f}%</td>'
                f'<td {_TD}>{atc:.1f}</td>'
                '</tr>'
            )
            rn += 1

    # ── Site Merch section ────────────────────────────────────────────
    if merch:
        m_atc = (sum(sf(r,'imp_m')*sf(r,'atc_rate') for r in merch)/m_imp) if m_imp else 0
        html += total_row('⬡ Site Merch Total', m_sov, m_ctr, m_atc, bl_merch)
        rn += 1
        for cn, rows in merch_cars.items():
            n = len(rows)
            for j, row in enumerate(rows):
                bg  = '#f8f9fa' if rn % 2 else '#ffffff'
                ctr = sf(row,'ctr'); imp = sf(row,'imp_m'); sov = sf(row,'sov_pct')
                cc  = '#16a34a' if ctr >= bl_merch else '#dc2626'
                if j == 0:
                    ctd = (
                        f'<td rowspan="{n}" style="padding:7px 10px;text-align:center;{_BD}'
                        f'background:#f0faf0;font-size:10px;font-weight:600;vertical-align:middle;">'
                        f'<span style="background:#22c55e;color:white;border-radius:3px;'
                        f'padding:1px 4px;font-size:9px;">Site Merch</span><br>'
                        f'<span style="font-size:9px;color:#555;">{cn[:22]}</span></td>'
                    )
                else:
                    ctd = ''
                atc = sf(row,'atc_rate')
                html += (
                    f'<tr style="background:{bg};">{ctd}'
                    f'<td {_TDL}>{row.get("display_label","—")}</td>'
                    f'<td {_TD}>{sov:.1f}%</td>'
                    f'<td style="padding:7px 12px;text-align:center;{_BD}font-weight:700;color:{cc};">{ctr:.2f}%</td>'
                    f'<td {_TD}>{atc:.1f}</td>'
                    '</tr>'
                )
                rn += 1

    # ── Grand total — no single benchmark applies; show both ──────────
    gt_imp = p_imp + m_imp
    gt_clk = sum(sf(r,'imp_m')*sf(r,'ctr')/100 for r in p13n+merch)
    gt_ctr = (gt_clk/gt_imp*100) if gt_imp else 0
    html += (
        f'<tr style="background:#f1f5f9;font-weight:800;font-size:13px;border-top:3px solid #041e42;">'
        f'<td colspan="2" style="padding:10px 14px;border:1px solid #cbd5e1;">GRAND TOTAL</td>'
        f'<td style="padding:10px 14px;text-align:center;border:1px solid #cbd5e1;">{gt_imp:.1f}M</td>'
        f'<td style="padding:10px 14px;text-align:center;border:1px solid #cbd5e1;">100%</td>'
        f'<td style="padding:10px 14px;text-align:center;border:1px solid #cbd5e1;">{bl:.2f}%</td>'
        f'<td style="padding:10px 14px;text-align:center;border:1px solid #cbd5e1;color:{gc};">{gt_ctr:.2f}%</td>'
        f'</tr>'
    )
    html += f'</tbody></table></div>{VAR_JS}'
    return html

def build_excel(rows, mod, bl, start, end, proj_map=None):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        thin = Side(border_style='thin', color='FFCCCCCC')
        bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
        wb   = openpyxl.Workbook(); ws = wb.active; ws.title = mod[:30]
        NAVY='FF041E42'; W='FFFFFFFF'; GRN='FF16a34a'; RED='FFdc2626'
        ws.merge_cells('A1:H1')
        ws['A1'] = f'HP Message Performance — {mod} | {start} to {end}'
        ws['A1'].font = Font(bold=True, size=14, color=W)
        ws['A1'].fill = PatternFill('solid', fgColor=NAVY)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A2:H2')
        ws['A2'] = f'FYTD CTR Baseline: {bl:.2f}% | Merch only | App iOS + Android | SOV Projected column is editable'
        ws['A2'].font = Font(size=10, color='FF555555')
        ws['A2'].alignment = Alignment(horizontal='center')
        headers = ['Group','Message / Asset','Impressions (M)','SOV %','SOV Projected (%)','SOV Variance','FYTD CTR Baseline','CTR %']
        col_w   = [30, 38, 16, 10, 18, 14, 18, 10]
        for i,(h,w) in enumerate(zip(headers, col_w), start=1):
            c=ws.cell(row=3,column=i,value=h)
            c.font=Font(bold=True,color=W,size=11)
            c.fill=PatternFill('solid',fgColor=NAVY)
            c.alignment=Alignment(horizontal='center',wrap_text=True)
            c.border=bdr
            ws.column_dimensions[chr(64+i)].width=w
        for rn, row in enumerate(rows, start=4):
            ctr = float(row.get('ctr') or 0)
            sov = float(row.get('sov_pct') or row.get('m0_sov') or 0)
            imt = row.get('_is_total', False)
            bg  = PatternFill('solid', fgColor=row.get('_bg','FFFFFFFF'))
            fn  = Font(bold=imt, size=11, color=row.get('_fg','FF000000'))
            vals= [row.get('_group',row.get('m0_group','')),
                   row.get('_label',row.get('message_name',row.get('display_label',''))),
                   round(float(row.get('imp_m') or 0),2),
                   f'{sov:.1f}%','','',f'{bl:.2f}%',f'{ctr:.2f}%']
            for ci,val in enumerate(vals,start=1):
                c=ws.cell(row=rn,column=ci,value=val)
                c.fill=bg; c.font=fn; c.border=bdr
                c.alignment=Alignment(horizontal='left' if ci<=2 else 'center',wrap_text=True)
                if ci==8: c.font=Font(bold=True,size=11,color=GRN if ctr>=bl else RED)
        buf=BytesIO(); wb.save(buf); buf.seek(0); return buf
    except Exception as e: print(f'Excel err: {e}'); return None


def flatten_for_excel(data, module):
    rows = []
    def sf(r,k): return float(r.get(k) or 0)
    if module == 'hpov':
        groups = OrderedDict()
        for r in data:
            g=r['m0_group']
            if g not in groups: groups[g]={'m0_sov':sf(r,'m0_sov'),'rows':[]}
            groups[g]['rows'].append(r)
        for gn,gd in groups.items():
            gi=sum(sf(r,'imp_m') for r in gd['rows'])
            gc=sum(sf(r,'imp_m')*sf(r,'ctr')/100 for r in gd['rows'])
            rows.append({'_group':gn,'_label':'','imp_m':gi,
                         'sov_pct':gd['m0_sov'],'ctr':(gc/gi*100) if gi else 0,
                         '_is_total':True,'_bg':'FFF0F4FF','_fg':'FF041E42'})
            for r in gd['rows']:
                rows.append({**r,'_group':'','_label':r.get('message_name',''),
                             '_is_total':False,'_bg':'FFFFFFFF'})
    else:
        p13n =[r for r in data if r.get('carousel_type')=='P13N']
        merch=[r for r in data if r.get('carousel_type')=='Site Merch']
        total_imp=sum(sf(r,'imp_m') for r in data)
        if p13n:
            pi=sum(sf(r,'imp_m') for r in p13n)
            pc=sum(sf(r,'imp_m')*sf(r,'ctr')/100 for r in p13n)
            bl_p13n_f = float(p13n[0].get('fytd_baseline_p13n') or 0) if p13n else 0
            rows.append({'_group':'P13N Total','_label':'','imp_m':pi,
                         'sov_pct':round(pi/total_imp*100,1) if total_imp else 0,
                         'ctr':(pc/pi*100) if pi else 0,'_is_total':True,
                         'fytd_baseline':bl_p13n_f,'_bg':'FFE8F0FC','_fg':'FF041E42'})
            for r in p13n:
                rows.append({**r,'_group':'','_label':r.get('display_label',''),
                             '_is_total':False,'fytd_baseline':bl_p13n_f,'_bg':'FFFFFFFF'})
        if merch:
            mi=sum(sf(r,'imp_m') for r in merch)
            mc=sum(sf(r,'imp_m')*sf(r,'ctr')/100 for r in merch)
            cars=OrderedDict()
            for r in merch:
                cn=r['carousel_nm']
                if cn not in cars: cars[cn]=[]
                cars[cn].append(r)
            rows.append({'_group':'Site Merch Total','_label':'','imp_m':mi,
                         'sov_pct':round(mi/total_imp*100,1) if total_imp else 0,
                         'ctr':(mc/mi*100) if mi else 0,'_is_total':True,'_bg':'FFF0FAF0','_fg':'FF041E42'})
            bl_merch_f = float(merch[0].get('fytd_baseline_merch') or 0) if merch else 0
            rows[-len(merch)-1]['fytd_baseline'] = bl_merch_f  # update Site Merch total row
            for cn,crows in cars.items():
                for r in crows:
                    rows.append({**r,'_group':cn,'_label':r.get('display_label',''),
                                 '_is_total':False,'fytd_baseline':bl_merch_f,'_bg':'FFFFFFFF'})
    return rows


@app.get('/debug')
async def debug_sig(
    start_date: str = Query(default='2026-04-13'),
    end_date:   str = Query(default='2026-05-09'),
    module:     str = Query(default='sig'),
    language:   str = Query(default='English'),
):
    """Debug endpoint — returns JSON with row count and first 3 rows."""
    try:
        if module == 'sig':
            data = get_sig_data(start_date, end_date, language)
        else:
            data = get_hpov_data(start_date, end_date, language)
        flat = flatten_for_excel(data, module)
        sample = flat[:3] if flat else []
        # Convert to serializable types
        def clean(r):
            return {k: (float(v) if hasattr(v,'__float__') and not isinstance(v,(str,bool)) else str(v) if not isinstance(v,(str,int,float,bool,type(None))) else v)
                    for k,v in r.items()}
        return JSONResponse({'module': module, 'total_rows': len(flat),
                             'data_rows': len(data), 'sample': [clean(r) for r in sample]})
    except Exception as e:
        import traceback
        return JSONResponse({'error': str(e), 'trace': traceback.format_exc()}, status_code=500)


@app.get('/download-excel')
async def download_excel(
    start_date:              str = Query(default='2026-04-13'),
    end_date:                str = Query(default='2026-05-09'),
    module:                  str = Query(default='hpov'),
    language:                str = Query(default='English'),
    selected_carousels:      str = Query(default=''),
    selected_messages:       str = Query(default=''),
    selected_hpov_messages:  str = Query(default=''),
    services_m0:             str = Query(default=''),
    projections:             str = Query(default=''),
):
    sc   = [c.strip() for c in selected_carousels.split('|')     if c.strip()]
    sm   = [m.strip() for m in selected_messages.split('|')      if m.strip()]
    shm  = [m.strip() for m in selected_hpov_messages.split('|') if m.strip()]
    svc  = [s.strip() for s in services_m0.split('|')            if s.strip()]
    proj_map = {}
    if projections:
        try: proj_map = json.loads(projections)
        except Exception as pe: print(f'[DL] proj parse err: {pe}')
    if module == 'hpov':
        data = get_hpov_data(start_date, end_date, language, svc or None)
        if shm:
            data = [r for r in data if r.get('message_name') in set(shm)]
        bl   = float(data[0].get('fytd_baseline',0.21)) if data else 0.21
        lbl  = 'HPOV'
    else:
        data = get_sig_data(start_date, end_date, language, sc or None, sm or None)
        print(f'[DL] SIG data rows: {len(data)}, sc={sc}, sm={sm}')
        bl   = float(data[0].get('fytd_baseline_merch', data[0].get('fytd_baseline', 1.11))) if data else 1.11
        lbl  = 'SIG'
    flat = flatten_for_excel(data, module)
    print(f'[DL] flat rows: {len(flat)}, bl={bl}')
    buf  = build_excel(flat, lbl, bl, start_date, end_date)
    if buf is None:
        from fastapi.responses import PlainTextResponse
        return PlainTextResponse('Excel build failed — check server log', status_code=500)
    fname = f'hp-{module}-{start_date}-to-{end_date}.xlsx'
    return StreamingResponse(buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{fname}"'})


@app.get('/', response_class=HTMLResponse)
async def home(
    start_date:         str = Query(default='2026-04-13'),
    end_date:           str = Query(default='2026-05-09'),
    module:             str = Query(default='hpov'),
    view:               str = Query(default='chart'),
    language:           str = Query(default='English'),
    selected_carousels:     str = Query(default=''),
    selected_messages:      str = Query(default=''),
    selected_hpov_messages: str = Query(default=''),
    services_m0:            str = Query(default=''),
):
    sc  = [c.strip() for c in selected_carousels.split('|')     if c.strip()]
    sm  = [m.strip() for m in selected_messages.split('|')      if m.strip()]
    shm = [m.strip() for m in selected_hpov_messages.split('|') if m.strip()]
    svc = [s.strip() for s in services_m0.split('|')            if s.strip()]
    sc_p  = '|'.join(sc);  sm_p = '|'.join(sm)
    shm_p = '|'.join(shm); svc_p = '|'.join(svc)

    extra_html = ''

    if module == 'hpov':
        # Services m0 selector
        all_m0   = get_hpov_m0_list(start_date, end_date, language)
        svc_set  = set(svc) if svc else set(DEFAULT_SERVICES.split('|'))
        data     = get_hpov_data(start_date, end_date, language, list(svc_set) if svc else None)
        baseline = float(data[0].get('fytd_baseline',0.21)) if data else 0.21
        mod_title = 'HPOV — AutoScroll Cards 1–5'
        mod_sub   = f'hp_module_name IN AutoScroll Card 1-5 | FYTD Baseline: {baseline:.2f}%'
        all_hpov_msgs = get_hpov_messages(start_date, end_date, language, list(svc_set) if svc else None)
        content   = render_hpov_chart(data) if view=='chart' else render_hpov_table(data, shm or None)

        # Build services selector UI
        m0_chks = ''
        for m0 in all_m0:
            nm  = m0['m0_nm']
            imp = round(float(m0.get('imp') or 0)/1e6, 1)
            chk = 'checked' if nm in svc_set else ''
            sf  = nm.replace('"','&quot;')
            m0_chks += (
                f'<label style="display:flex;align-items:center;gap:6px;font-size:12px;'
                f'padding:3px 0;cursor:pointer;">'
                f'<input type="checkbox" name="svc_opt" value="{sf}" {chk}>'
                f' {nm} <span style="color:#888;font-size:10px;">({imp}M)</span>'
                f'</label>'
            )
        # Build services dropdown options
        svc_opts = []
        for m0 in all_m0:
            nm  = m0['m0_nm']
            mi  = round(float(m0.get('imp') or 0)/1e6, 1)
            svc_opts.append((nm, nm, mi))
        svc_dd = _dd_panel('svc', '⚙️ Services Group', svc_opts, svc_set, 'ddApply', grouped=False)
        extra_html = (
            f'<div style="display:flex;align-items:center;flex-wrap:wrap;gap:8px;'
            f'background:#fff8ed;border:1px solid #fed7aa;border-radius:8px;'
            f'padding:12px 16px;margin-bottom:10px;">'
            f'<span style="font-size:12px;font-weight:700;color:#9a3412;white-space:nowrap;">'
            f'⚙️ Group as "Services":</span>'
            + svc_dd +
            f'<input type="hidden" id="svcHidden" name="services_m0" value="{svc_p}">'
            f'</div>'
        )

        # HPOV message filter dropdown
        if all_hpov_msgs:
            msg_opts_grouped = []
            for m in all_hpov_msgs:
                mn = m['message_name']
                mi = round(float(m.get('imp') or 0)/1e6, 1)
                msg_opts_grouped.append((m['m0_group'], mn, mn, mi))
            msg_dd = _dd_panel('hpovmsg', '🔍 Message Filter', msg_opts_grouped, set(shm) if shm else None, 'ddApply', grouped=True)
            extra_html += (
                f'<div style="display:flex;align-items:center;flex-wrap:wrap;gap:8px;'
                f'background:#f0f9ff;border:1px solid #bae6fd;border-radius:8px;'
                f'padding:12px 16px;margin-top:10px;">'
                f'<span style="font-size:12px;font-weight:700;color:#0369a1;white-space:nowrap;">'
                f'🔍 Messages:</span>'
                + msg_dd +
                f'<input type="hidden" id="hpovMsgHid" name="selected_hpov_messages" value="{shm_p}">'
                f'</div>'
            )

    else:  # SIG
        all_cars = get_sig_carousels(start_date, end_date, language)
        all_msgs = get_sig_messages(start_date, end_date, language, sc or None)
        data     = get_sig_data(start_date, end_date, language, sc or None, sm or None)
        bl_p13n_h  = float(data[0].get('fytd_baseline_p13n',  0)) if data else 0
        bl_merch_h = float(data[0].get('fytd_baseline_merch', 0)) if data else 0
        baseline   = bl_merch_h  # use merch as fallback for page-level
        mod_title  = 'SIG — Scrollable Item Grid'
        mod_sub    = (f'SIG Cards 1–6 | FYTD: '
                      f'P13N {bl_p13n_h:.2f}% | Site Merch {bl_merch_h:.2f}%')
        content   = render_sig_chart(data) if view=='chart' else render_sig_table(data, sm or None)

        # Carousel filter
        car_chks = ''
        for car in all_cars:
            cn  = car['carousel_nm']; imp = round(float(car.get('imp') or 0)/1e6,1)
            chk = 'checked' if (not sc or cn in sc) else ''
            sf2 = cn.replace('"','&quot;')
            car_chks += (
                f'<label style="display:flex;align-items:center;gap:6px;font-size:12px;padding:3px 0;cursor:pointer;">'
                f'<input type="checkbox" name="car_opt" value="{sf2}" {chk}> {cn} <span style="color:#888;font-size:10px;">({imp}M)</span></label>'
            )
        # Message filter
        msg_chks = ''
        for msg in all_msgs:
            ml  = msg['display_label']; mi = round(float(msg.get('imp') or 0)/1e6,1)
            chk = 'checked' if (not sm or ml in sm) else ''
            ms  = ml.replace('"','&quot;')
            msg_chks += (
                f'<label style="display:flex;align-items:center;gap:6px;font-size:12px;padding:3px 0;cursor:pointer;">'
                f'<input type="checkbox" name="msg_opt" value="{ms}" {chk}> {ml} <span style="color:#888;font-size:10px;">({mi}M)</span></label>'
            )
        # Build SIG dropdowns
        car_opts = [(c['carousel_nm'], c['carousel_nm'], round(float(c.get('imp') or 0)/1e6,1)) for c in all_cars]
        msg_opts = [(m['display_label'], m['display_label'], round(float(m.get('imp') or 0)/1e6,1)) for m in all_msgs]
        car_dd = _dd_panel('sigcar', '🎠 Carousels', car_opts, set(sc) if sc else None, 'ddApply', grouped=False)
        msg_dd = _dd_panel('sigmsg', '💬 Messages', msg_opts, set(sm) if sm else None, 'ddApply', grouped=False)
        extra_html = (
            f'<div style="display:flex;align-items:center;flex-wrap:wrap;gap:8px;'
            f'background:#f8f9fa;border:1px solid #e5e7eb;border-radius:8px;'
            f'padding:12px 16px;margin-bottom:10px;">'
            f'<span style="font-size:12px;font-weight:700;color:#041e42;white-space:nowrap;">'
            f'SIG Filters:</span>'
            + car_dd
            + msg_dd +
            f'<input type="hidden" id="carHid" name="selected_carousels" value="{sc_p}">'
            f'<input type="hidden" id="msgHid" name="selected_messages" value="{sm_p}">'
            f'</div>'
        )

    # Tab styles
    _A='background:#041e42;color:white;border-radius:6px 6px 0 0;'
    _I='background:#e5e7eb;color:#374151;border-radius:6px 6px 0 0;'
    _VA='background:#0071CE;color:white;border-radius:6px;'
    _VI='background:#e5e7eb;color:#374151;border-radius:6px;'
    th = _A if module=='hpov' else _I
    ts = _A if module=='sig'  else _I
    vc = _VA if view=='chart' else _VI
    vt = _VA if view=='table' else _VI

    lang_opts = ''.join(f'<option value="{v}" {"selected" if v==language else ""} >{v}</option>' for v in ['English','Spanish','All'])

    xl_url = ('/download-excel'
              f'?start_date={start_date}&end_date={end_date}'
              f'&module={module}&language={quote(language)}'
              f'&selected_carousels={quote(sc_p)}'
              f'&selected_messages={quote(sm_p)}'
              f'&selected_hpov_messages={quote(shm_p)}'
              f'&services_m0={quote(svc_p)}')

    chart_scripts = (
        '<script src="https://cdn.jsdelivr.net/npm/chart.js@4"></script>'
        '<script src="https://cdn.jsdelivr.net/npm/chartjs-plugin-datalabels@2"></script>'
        '<script src="https://cdn.jsdelivr.net/npm/chartjs-plugin-annotation@2"></script>'
    ) if view == 'chart' else ''

    html = (
        '<!DOCTYPE html><html lang="en"><head>'
        '<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">'
        '<title>HP Message Performance</title>'
        + chart_scripts +
        '<style>'
        'body{font-family:\'Segoe UI\',Arial,sans-serif;background:#f3f4f6;margin:0;padding:20px;}'
        '.card{background:white;border-radius:12px;box-shadow:0 2px 10px rgba(0,0,0,.1);'
        'padding:24px;max-width:1700px;margin:0 auto;}'
        'a{text-decoration:none;}'
        '</style></head><body><div class="card">'
        '<div style="margin-bottom:18px;">'
        '<div style="font-size:22px;font-weight:800;color:#041e42;">📊 HP Message Performance</div>'
        '<div style="font-size:11px;color:#888;margin-top:4px;">'
        'FYTD Baseline: Jan 31 2026 → latest &nbsp;|&nbsp; Merch only &nbsp;|&nbsp; App iOS + Android'
        '</div></div>'
        f'<form method="get" class="mf" '
        f'style="display:flex;align-items:flex-end;gap:14px;flex-wrap:wrap;'
        f'background:#f8f9fa;border:1px solid #e5e7eb;border-radius:8px;'
        f'padding:14px 18px;margin-bottom:16px;">'
        f'<input type="hidden" name="module" value="{module}">'
        f'<input type="hidden" name="view" value="{view}">'
        f'<input type="hidden" name="services_m0" id="svcHidden" value="{svc_p}">'
        f'<input type="hidden" name="selected_carousels" id="carHid" value="{sc_p}">'
        f'<input type="hidden" name="selected_messages" id="msgHid" value="{sm_p}">'
        f'<input type="hidden" name="selected_hpov_messages" id="hpovMsgHid" value="{shm_p}">'
        f'<div><label style="font-size:12px;font-weight:600;display:block;margin-bottom:4px;">Start Date</label>'
        f'<input type="date" name="start_date" value="{start_date}" style="border:1px solid #d1d5db;border-radius:6px;padding:7px 10px;font-size:13px;"></div>'
        f'<div><label style="font-size:12px;font-weight:600;display:block;margin-bottom:4px;">End Date</label>'
        f'<input type="date" name="end_date" value="{end_date}" style="border:1px solid #d1d5db;border-radius:6px;padding:7px 10px;font-size:13px;"></div>'
        f'<div><label style="font-size:12px;font-weight:600;display:block;margin-bottom:4px;">Language</label>'
        f'<select name="language" style="border:1px solid #d1d5db;border-radius:6px;padding:7px 10px;font-size:13px;">{lang_opts}</select></div>'
        f'<button type="submit" style="background:#041e42;color:white;border:none;border-radius:6px;padding:8px 20px;font-size:13px;cursor:pointer;font-weight:600;">Generate</button>'
        f'</form>'
        f'<div style="display:flex;gap:0;border-bottom:2px solid #041e42;margin-bottom:0;">'
        f'<a href="?start_date={start_date}&end_date={end_date}&module=hpov&view={view}&language={language}" style="padding:10px 28px;font-size:14px;font-weight:700;{th}">📈 HPOV</a>'
        f'<a href="?start_date={start_date}&end_date={end_date}&module=sig&view={view}&language={language}" style="padding:10px 28px;font-size:14px;font-weight:700;{ts}">📋 SIG</a>'
        f'</div>'
        f'<div style="display:flex;align-items:center;justify-content:space-between;padding:14px 0 12px;flex-wrap:wrap;gap:10px;">'
        f'<div><div style="font-size:16px;font-weight:700;color:#041e42;">{mod_title}</div>'
        f'<div style="font-size:11px;color:#888;">{mod_sub}</div></div>'
        f'<div style="display:flex;gap:8px;align-items:center;">'
        f'<a href="?start_date={start_date}&end_date={end_date}&module={module}&view=chart&language={language}&selected_carousels={sc_p}&selected_messages={sm_p}&services_m0={svc_p}" style="padding:7px 20px;font-size:13px;font-weight:600;{vc}">📈 Chart</a>'
        f'<a href="?start_date={start_date}&end_date={end_date}&module={module}&view=table&language={language}&selected_carousels={sc_p}&selected_messages={sm_p}&services_m0={svc_p}" style="padding:7px 20px;font-size:13px;font-weight:600;{vt}">📋 Table</a>'
        f'<button type="button" id="xlBtn" onclick="kDlExcel()" '
        f'style="padding:7px 20px;font-size:13px;font-weight:600;background:#22c55e;'
        f'color:white;border-radius:6px;border:none;cursor:pointer;">⬇ Excel</button>'
        f'<script>'
        f'var _XL_BASE = "{xl_url}";'
        f'function kDlExcel(){{'
        f'  var proj={{}};'
        f'  document.querySelectorAll("input[data-msg]").forEach(function(inp){{'
        f'    var v=parseFloat(inp.value);'
        f'    if(!isNaN(v) && inp.dataset.msg) proj[inp.dataset.msg]=v;'
        f'  }});'
        f'  var url=_XL_BASE+(Object.keys(proj).length>0?"&projections="+encodeURIComponent(JSON.stringify(proj)):"");'
        f'  window.location.href=url;'
        f'}}'
        f'</script>'
        f'</div></div>'
        + """

<style>
.dd-wrap{position:relative;display:inline-block;margin-right:12px;margin-bottom:8px;}
.dd-btn{background:white;border:1px solid #d1d5db;border-radius:6px;padding:7px 14px;
  font-size:13px;cursor:pointer;display:flex;align-items:center;gap:8px;min-width:200px;
  justify-content:space-between;font-weight:600;color:#374151;}
.dd-btn:hover{border-color:#6b7280;}
.dd-panel{position:absolute;top:calc(100% + 4px);left:0;z-index:999;background:white;
  border:1px solid #d1d5db;border-radius:8px;box-shadow:0 8px 24px rgba(0,0,0,.15);
  min-width:300px;max-width:380px;padding:10px;}
.dd-search{width:100%;border:1px solid #e5e7eb;border-radius:5px;padding:6px 10px;
  font-size:12px;margin-bottom:6px;box-sizing:border-box;}
.dd-opts{max-height:280px;overflow-y:auto;padding:2px 0;}
.dd-opts label{display:flex;align-items:center;gap:7px;font-size:12px;padding:4px 6px;
  cursor:pointer;border-radius:4px;color:#374151;}
.dd-opts label:hover{background:#f3f4f6;}
.dd-opts .dd-grp{font-size:11px;font-weight:700;color:#6b7280;padding:6px 6px 2px;
  border-top:1px solid #f0f0f0;margin-top:2px;}
.dd-footer{display:flex;gap:6px;margin-top:8px;border-top:1px solid #f0f0f0;padding-top:8px;}
.dd-apply{flex:1;background:#041e42;color:white;border:none;border-radius:5px;
  padding:6px 12px;font-size:12px;cursor:pointer;font-weight:600;}
.dd-apply:hover{background:#0a3161;}
.dd-clear{background:#f3f4f6;color:#374151;border:none;border-radius:5px;
  padding:6px 10px;font-size:12px;cursor:pointer;}
.dd-selall{background:#f3f4f6;color:#374151;border:none;border-radius:5px;
  padding:6px 10px;font-size:12px;cursor:pointer;}
</style>
<script>
function ddToggle(id){
  var p=document.getElementById(\'ddp_\'+id);
  var open=p.style.display!==\'none\';
  document.querySelectorAll(\'.dd-panel\').forEach(function(el){el.style.display=\'none\';});
  if(!open) p.style.display=\'block\';
}
function ddSearch(id,q){
  var opts=document.querySelectorAll(\'#ddp_\'+id+\' .dd-opts label:not(.dd-grp)\');
  opts.forEach(function(l){
    l.style.display=l.textContent.toLowerCase().includes(q.toLowerCase())?\'\':\'none\';
  });
}
function ddSelAll(id,chk){
  document.querySelectorAll(\'#ddp_\'+id+\' input[type=checkbox]\').forEach(function(c){c.checked=chk;});
}
document.addEventListener(\'click\',function(e){
  if(!e.target.closest(\'.dd-wrap\'))
    document.querySelectorAll(\'.dd-panel\').forEach(function(el){el.style.display=\'none\';});
});
</script>

        """
        + f'<script>'
        + f'function ddApply(id){{'
        + f'var v=Array.from(document.querySelectorAll(".dd-opt-"+id+":checked")).map(c=>c.value).join("|");'
        + f'if(id==="svc"){{document.getElementById("svcHidden").value=v;}}'
        + f'else if(id==="hpovmsg"){{document.getElementById("hpovMsgHid").value=v;}}'
        + f'else if(id==="sigcar"){{document.getElementById("carHid").value=v;document.getElementById("msgHid").value="";}}'
        + f'else if(id==="sigmsg"){{document.getElementById("msgHid").value=v;}}'
        + f'document.querySelector("form.mf").submit();}}'
        + f'</script>'
        + extra_html
        + content +
        '</div></body></html>'
    )
    return HTMLResponse(content=html)


if __name__ == '__main__':
    import uvicorn
    uvicorn.run(app, host='127.0.0.1', port=8002)
